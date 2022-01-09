#NAME:EKOUO HONORE MOUKOUELLE
#MATRICULE: ICTU20201251

from openpyxl import load_workbook
import pandas as pd

def xlsx():# this function will will in charge of .xlsx file
    wb=load_workbook('employeedata.xlsx')
    ws=wb.active
    sheet = wb['Feuil1']
    
    for i in range(5,sheet.max_row+1):
        cell=sheet.cell(i,3)
        if 'helpinghands.cm' in cell.value:
            update=(cell.value).replace('helpinghands.cm','handsinhands.org')
            sheet.cell(i,3).value=update
        
        wb.save('updated_emails.xlsx')
        
    
def csv(): # this function will will in charge of .cvs file
# reading the csv file
    df = pd.read_csv("employeedata.csv")
    
    # updating the column value/data
    df['ADDRESS'] = df['ADDRESS'].replace({'@helpinghands.cm':'@handsinhands.org'})
    
    # writing into the file
    df.to_csv("updated_email.csv", index=False)
    
    print(df)
        
xlsx() #Here we update the .xlsx file i.e employeedata.xlsx --> updated_emails.xlsx
csv()  #Here we update the .csv file i.e employeedata.csv --> updated_emails.csv